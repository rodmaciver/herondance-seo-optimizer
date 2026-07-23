"""Tests for the Change Ledger: row construction and xlsx append."""
import io
import unittest

import openpyxl

from batch_runner import _ledger_new_url, _ledger_row, _plan_decision
from src import sheets_client
from src.schema import ExecutionPlan, PageSnapshot, PlanItem


def _make_plan(**kwargs) -> ExecutionPlan:
    defaults = dict(
        page_url="https://herondance.org/old-page",
        primary_keyword="wu wei",
        secondary_keywords=[],
        items=[
            PlanItem(field="seo_title", decision="Wu Wei: The Art of Effortless Flow"),
            PlanItem(field="meta_description", decision="A quiet reflection on wu wei."),
        ],
        body_changes=[],
        redirect_mapping=None,
    )
    defaults.update(kwargs)
    return ExecutionPlan(**defaults)


def _make_snapshot() -> PageSnapshot:
    return PageSnapshot(
        url="https://herondance.org/old-page",
        title="Old Title",
        meta_description="Old description.",
    )


def _empty_queue_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheets_client.QUEUE_TAB
    ws.append(["Top pages"])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


class PlanDecisionTests(unittest.TestCase):
    def test_returns_decision_for_field(self):
        plan = _make_plan()
        self.assertEqual(_plan_decision(plan, "seo_title"), "Wu Wei: The Art of Effortless Flow")

    def test_missing_field_returns_empty(self):
        plan = _make_plan()
        self.assertEqual(_plan_decision(plan, "h1"), "")


class LedgerNewUrlTests(unittest.TestCase):
    def test_no_redirect_keeps_url(self):
        plan = _make_plan()
        url = "https://herondance.org/old-page"
        self.assertEqual(_ledger_new_url(url, plan), url)

    def test_redirect_yields_new_url_on_same_domain(self):
        plan = _make_plan(redirect_mapping="/old-page -> /wu-wei-effortless-flow 301")
        self.assertEqual(
            _ledger_new_url("https://herondance.org/old-page", plan),
            "https://herondance.org/wu-wei-effortless-flow",
        )


class LedgerRowTests(unittest.TestCase):
    def test_row_has_one_value_per_header_column(self):
        row = _ledger_row("https://herondance.org/old-page", _make_snapshot(), _make_plan())
        self.assertEqual(len(row), len(sheets_client.LEDGER_HEADER))

    def test_row_captures_old_and_new_values(self):
        row = _ledger_row("https://herondance.org/old-page", _make_snapshot(), _make_plan())
        header = sheets_client.LEDGER_HEADER
        as_dict = dict(zip(header, row))
        self.assertEqual(as_dict["old_title"], "Old Title")
        self.assertEqual(as_dict["new_title"], "Wu Wei: The Art of Effortless Flow")
        self.assertEqual(as_dict["old_meta_description"], "Old description.")
        self.assertEqual(as_dict["new_meta_description"], "A quiet reflection on wu wei.")
        self.assertEqual(as_dict["primary_keyword"], "wu wei")
        self.assertEqual(as_dict["implemented_date"], "")


class AppendLedgerRowTests(unittest.TestCase):
    def test_creates_tab_with_header_then_appends(self):
        xlsx = _empty_queue_xlsx()
        row = _ledger_row("https://herondance.org/old-page", _make_snapshot(), _make_plan())
        xlsx = sheets_client.append_ledger_row(xlsx, row)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        self.assertIn(sheets_client.LEDGER_TAB, wb.sheetnames)
        ws = wb[sheets_client.LEDGER_TAB]
        self.assertEqual(
            [c.value for c in ws[1]], sheets_client.LEDGER_HEADER
        )
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=2).value, "https://herondance.org/old-page")

    def test_second_append_does_not_duplicate_header(self):
        xlsx = _empty_queue_xlsx()
        row = _ledger_row("https://herondance.org/old-page", _make_snapshot(), _make_plan())
        xlsx = sheets_client.append_ledger_row(xlsx, row)
        xlsx = sheets_client.append_ledger_row(xlsx, row)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb[sheets_client.LEDGER_TAB]
        self.assertEqual(ws.max_row, 3)  # header + two rows

    def test_queue_tab_untouched(self):
        xlsx = _empty_queue_xlsx()
        row = _ledger_row("https://herondance.org/old-page", _make_snapshot(), _make_plan())
        xlsx = sheets_client.append_ledger_row(xlsx, row)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws = wb[sheets_client.QUEUE_TAB]
        self.assertEqual(ws.cell(row=1, column=1).value, "Top pages")
        self.assertEqual(ws.max_row, 1)


if __name__ == "__main__":
    unittest.main()


class NoMaterialChangesBannerTests(unittest.TestCase):
    def _all_keep_plan(self):
        return _make_plan(
            items=[
                PlanItem(field="seo_title", decision="keep current"),
                PlanItem(field="meta_description", decision="keep current"),
                PlanItem(field="url_slug", decision="keep current"),
                PlanItem(field="h1", decision="keep current"),
            ],
            body_changes=[],
        )

    def test_banner_present_when_all_keep_current(self):
        from src.pastepack import render_pastepack
        result = render_pastepack(self._all_keep_plan(), _make_snapshot(), operator="test")
        self.assertIn("NO CHANGES RECOMMENDED", result["markdown"])

    def test_banner_absent_when_title_changes(self):
        from src.pastepack import render_pastepack
        plan = self._all_keep_plan()
        plan.items[0].decision = "A Genuinely New Title"
        result = render_pastepack(plan, _make_snapshot(), operator="test")
        self.assertNotIn("NO CHANGES RECOMMENDED", result["markdown"])


class AdsStrengthTests(unittest.TestCase):
    def _plan_with_pool(self, pool):
        from src.schema import EvaluatedKeyword
        return _make_plan(keyword_pool=[EvaluatedKeyword(**kw) for kw in pool])

    def test_sums_selected_keyword_volume(self):
        from batch_runner import _ads_strength
        plan = self._plan_with_pool([
            {"term": "wu wei", "volume": 5400, "has_data": True, "selected": True},
            {"term": "effortless action", "volume": 900, "has_data": True, "selected": True},
            {"term": "quiet bird", "volume": 20, "has_data": True, "selected": False},
        ])
        self.assertEqual(_ads_strength(plan), 6300)

    def test_falls_back_to_top_pool_volumes_when_selected_have_none(self):
        from batch_runner import _ads_strength
        plan = self._plan_with_pool([
            {"term": "a", "volume": None, "selected": True},
            {"term": "b", "volume": 100, "selected": False},
            {"term": "c", "volume": 50, "selected": False},
            {"term": "d", "volume": 10, "selected": False},
            {"term": "e", "volume": 5, "selected": False},
        ])
        self.assertEqual(_ads_strength(plan), 160)

    def test_empty_pool_scores_zero(self):
        from batch_runner import _ads_strength
        self.assertEqual(_ads_strength(_make_plan()), 0)

    def test_batch_sorts_strongest_first(self):
        entries = [
            {"url": "weak", "strength_score": 30},
            {"url": "strong", "strength_score": 6300},
            {"url": "middle", "strength_score": 900},
        ]
        entries.sort(key=lambda e: e.get("strength_score", 0), reverse=True)
        self.assertEqual([e["url"] for e in entries], ["strong", "middle", "weak"])
