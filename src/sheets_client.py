"""Google Drive integration: read SEO queue xlsx and upload paste-pack files."""
from __future__ import annotations

import io
import json
import os
from pathlib import Path

import pandas as pd

DRIVE_FILE_ID = "1LIa3h_oksQUVtjHCQE5nRPz7kZko361Z"
QUEUE_TAB = "URLs to Do"
SHARED_DRIVE_FOLDER_ID = "0AHp_CfZhbhV0Uk9PVA"
SCOPES = ["https://www.googleapis.com/auth/drive"]

# Local dev fallback: place a service account key here (never commit it).
_LOCAL_KEY_PATH = Path(__file__).resolve().parent.parent / "config" / "sheets-key.json"


def _credentials():
    from google.oauth2 import service_account

    key_json = os.environ.get("SHEETS_SERVICE_ACCOUNT_KEY")
    if key_json:
        info = json.loads(key_json)
        return service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
    if _LOCAL_KEY_PATH.exists():
        return service_account.Credentials.from_service_account_file(
            str(_LOCAL_KEY_PATH), scopes=SCOPES
        )
    raise RuntimeError(
        "No Google Drive credentials found. Set SHEETS_SERVICE_ACCOUNT_KEY env var "
        "or place a service account key at config/sheets-key.json (never commit it)."
    )


def available() -> bool:
    """True if credentials are present — used to decide whether to use Drive or local xlsx."""
    return bool(os.environ.get("SHEETS_SERVICE_ACCOUNT_KEY")) or _LOCAL_KEY_PATH.exists()


def _download_queue_bytes() -> bytes:
    """Download the queue xlsx from Google Drive and return raw bytes."""
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    creds = _credentials()
    service = build("drive", "v3", credentials=creds, cache_discovery=False)

    request = service.files().get_media(fileId=DRIVE_FILE_ID)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()

    return buffer.getvalue()


def read_queue() -> pd.DataFrame:
    """Download the xlsx from Google Drive and return a raw headerless DataFrame."""
    raw_bytes = _download_queue_bytes()
    return pd.read_excel(io.BytesIO(raw_bytes), sheet_name=QUEUE_TAB, header=None)


def read_queue_with_bytes() -> tuple[pd.DataFrame, bytes]:
    """Download the xlsx once; return (raw_headerless_df, raw_bytes).

    The raw bytes are handed to the batch runner so it can call
    update_queue_cell() / upload_queue() without a second download.
    """
    raw_bytes = _download_queue_bytes()
    df = pd.read_excel(io.BytesIO(raw_bytes), sheet_name=QUEUE_TAB, header=None)
    return df, raw_bytes


def update_queue_cell(xlsx_bytes: bytes, row: int, col: int, value: str) -> bytes:
    """Write a single cell in the in-memory xlsx and return the updated bytes.

    row / col are 1-based (openpyxl convention).
    """
    import openpyxl

    buf = io.BytesIO(xlsx_bytes)
    wb = openpyxl.load_workbook(buf)
    ws = wb[QUEUE_TAB]
    ws.cell(row=row, column=col, value=value)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


LEDGER_TAB = "Change Ledger"
LEDGER_HEADER = [
    "date", "old_url", "new_url", "old_title", "new_title",
    "old_meta_description", "new_meta_description", "primary_keyword",
    "implemented_date",
]


def append_ledger_row(xlsx_bytes: bytes, values: list) -> bytes:
    """Append one row to the Change Ledger tab (created with a header row if
    absent) and return the updated bytes.

    The ledger records what each optimization changed (titles, meta
    descriptions, URLs) so that before/after Search Console comparisons and
    change attribution are possible months later. The implemented_date column
    is left blank for the VA to fill in when the change goes live on
    Squarespace.
    """
    import openpyxl

    buf = io.BytesIO(xlsx_bytes)
    wb = openpyxl.load_workbook(buf)
    if LEDGER_TAB in wb.sheetnames:
        ws = wb[LEDGER_TAB]
    else:
        ws = wb.create_sheet(LEDGER_TAB)
        ws.append(LEDGER_HEADER)
    padded = list(values)[: len(LEDGER_HEADER)]
    padded += [""] * (len(LEDGER_HEADER) - len(padded))
    ws.append(padded)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def upload_queue(xlsx_bytes: bytes) -> None:
    """Overwrite the queue xlsx on Google Drive with updated bytes."""
    import os
    import tempfile

    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    creds = _credentials()
    service = build("drive", "v3", credentials=creds, cache_discovery=False)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(xlsx_bytes)
        tmp_path = tmp.name

    try:
        service.files().update(
            fileId=DRIVE_FILE_ID,
            media_body=MediaFileUpload(
                tmp_path,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            supportsAllDrives=True,
        ).execute()
    finally:
        os.unlink(tmp_path)


def upload_file(local_path: str, filename: str) -> str:
    """Upload any file to the Shared Drive folder, inferring mime type from extension.

    Returns the web view URL of the uploaded file.
    """
    import mimetypes

    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    mime_type, _ = mimetypes.guess_type(filename)
    if not mime_type:
        mime_type = "application/octet-stream"

    creds = _credentials()
    service = build("drive", "v3", credentials=creds, cache_discovery=False)

    uploaded = service.files().create(
        body={"name": filename, "parents": [SHARED_DRIVE_FOLDER_ID]},
        media_body=MediaFileUpload(local_path, mimetype=mime_type),
        supportsAllDrives=True,
        fields="id,webViewLink",
    ).execute()

    return uploaded.get("webViewLink", "")


def upload_docx(local_path: str, filename: str) -> str:
    """Upload a .docx file to the Shared Drive folder.

    Returns the web view URL of the uploaded file.
    supportsAllDrives=True is required for Shared Drive access.
    """
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    creds = _credentials()
    service = build("drive", "v3", credentials=creds, cache_discovery=False)

    file_metadata = {
        "name": filename,
        "parents": [SHARED_DRIVE_FOLDER_ID],
    }
    media = MediaFileUpload(
        local_path,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )
    uploaded = service.files().create(
        body=file_metadata,
        media_body=media,
        supportsAllDrives=True,
        fields="id,webViewLink",
    ).execute()

    return uploaded.get("webViewLink", "")

